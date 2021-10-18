'''
smart attendence
version: 5.1

Copyright © 2021  - by vip 
'''

#   ''pip install -r require_smart.txt'' 
import os
import xlrd
import xlwt
import getpass
import datetime
import openpyxl
import colorama
import mysql.connector as connector
from xlutils.copy import copy
from xlrd import open_workbook
from prettytable import PrettyTable,from_db_cursor
from colorama import Fore,Back,Style
from openpyxl.styles import PatternFill
colorama.init()
pt=PrettyTable()

def info():	
	print(' This will accept excel file which are present only in the folder of this executable file of python\n')
	print(' It take attendence in excel and in mysql also from excel \n')
	print(' You can use file excel file with extenstions -\n  .xls ,  .xlsx  ,  .xlsm  ,  .xlts  ,  .xltm \n')

today=datetime.datetime.now()
print(Fore.YELLOW + ' DATE-',today.strftime("%x"))
print(' TIME-',today.strftime("%X"))

time="{}{}_{}_{}_{}".format(today.strftime("%b").upper(),today.strftime("%d"),today.strftime("%Y"),today.strftime("%H"),today.strftime("%M"))

def co(host,usr,passwd,d=''):
	global con
	con=connector.connect(host=host,port='3306',user=usr,password=passwd,database=d)

print(Fore.RESET,Style.BRIGHT,Fore.MAGENTA)

#Print results from database in table
def print_t(cursor_object):
    table = from_db_cursor(cursor_object)
    table.align = 'l'
    print(table)
user=(input(Fore.YELLOW+'Enter Name: '))
if user=='':
	user = getpass.getuser()
	
print('\n****************************************************************')

print(Fore.CYAN)
pt.field_names=[ 'HELLO',user.upper(),'SIR']
pt.add_row([' ','GOOD MORNING , WELCOME TO SMART ATTENDENCE',' '])
print(pt)

# it take file name if any error found it again ask for file
print(Fore.RESET)
print(Fore.GREEN)

info()

d=''

while True:
	loc =str(input(Fore.MAGENTA + "Enter your file name: "))
	try:
		print(Fore.CYAN)
		#this try to open xls file of excel with the prettytable
		p=PrettyTable()
		wb = xlrd.open_workbook(loc)
		sheet =wb.sheet_by_index(0)
		sheet.cell_value(0,0)
		x=sheet.nrows
		p.field_names=sheet.row_values(0)
		for i in range(1,x):
			p.add_row(sheet.row_values(i))
		print(p)
		print(' ')
		y=int(input(Fore.MAGENTA + 'Enter column no. of Names: '))
		c=sheet.col_values(y-1)
		col=sheet.col_values(0)
		break
	except:
			print(Fore.CYAN)
			try:
				#this try to read your xlsx,xlsm,xltx and xltm files
				t=list()
				c=list()
				col=list()
				wb_obj = openpyxl.load_workbook(loc)
				sheet_obj = wb_obj.active
				m_col =sheet_obj.max_column+1
				m_row=sheet_obj.max_row
				
				def iter_rows(sheet_obj,a):
					result=list()
					for row in sheet_obj.iter_rows(a,a):
						for cell in row:
							result.append(cell.value)
							yield result
				z=list(iter_rows(sheet_obj,1))
				tab=PrettyTable()
				tab.field_names=z[0]
				no=m_row+1				
				for i in range(2,no):
					z=list(iter_rows(sheet_obj,i))
					tab.add_row(z[0])
				print(tab)
								
				def path(path,y,nam):
					wb_obj = openpyxl.load_workbook(path)
					sheet_obj = wb_obj.active
					for i in range(1, m_col):
						cell_obj = sheet_obj.cell(row = y, column = i)
						t.append(cell_obj.value)
					for i in range(1,m_row+1):
						cell_obj1=sheet_obj.cell(row=i,column=nam)
						c.append(cell_obj1.value)
				row1=int(input(Fore.MAGENTA + 'Enter column no of Roll no: '))
				nam=int(input(Fore.MAGENTA + 'Enter column no. of Names: '))

				for i in range(1,m_col):
					path(loc,i,nam)
					for i in range(1,m_row+1):
						cell_obj2=sheet_obj.cell(row=i,column=row1)
						col.append(cell_obj2.value)
						x=len(col)-m_row
						while x>0:
							col.pop(0)
							c.pop(0)
							x=x-1								
				
				
				break
			except:
				print(Fore.RED + "file is not found try again")
				print(' ')
print(Fore.GREEN)
#this function try to make letter into title form so that we can easily compared				
def cap(c):
		x=len(c)
		for i in range(x):
			a=c[i].title()
			c.append(a)
		while x>0:
			c.pop(0)
			x=x-1
			
cap(c)
'''
print(' ')
print(c[0],end='- ')
print(c[1:])
print(' ')
print(col[0],end='- ')
print(col[1:])
print(' ')
'''
l=list()

count=0

enter='Enter your Name or Roll no: '
print('1: Attendence in excel')
print('2: Attendence in My Sql')
print('3: Attendence in python(will not save)')
Q=int(input(Fore.MAGENTA + 'Your Input: '))

print(' ')

def ex():
	print('Exiting ....')
	print('Bye ...')
	eval('exit()')

if Q==1:
	print(' ')
	enter='Enter your Roll no: '
	print(' ')
	file=str(input(Fore.MAGENTA + 'Enter your excel file name with or .xlsx to save it: '))
	try:
	       	excel_file = openpyxl.load_workbook(loc)
	       	excel_sheet = excel_file.active
	       	excel_sheet.cell(row=1, column=m_col).value ='Attendence'
	       	for i in range(1,m_row):
	       		excel_sheet.cell(row=i+1, column=m_col).value ='Absent'
	       		excel_sheet.cell(row=i+1, column=m_col).fill=PatternFill('solid',fgColor='DC143C')
	       		excel_file.save(file)

	except:
		rb = open_workbook(loc)
		wb = copy(rb)
		s = wb.get_sheet(0)
		s.write(0,x-1,'Attendence')
		for i in range(0,x-1):
			s.write(i+1,x-1,'absent')
			wb.save(file)

elif Q==2:
	os.system('clear')
	print('Mysql connection making.....\n')
	for _ in range(3):
	   		_+=1
	   		print('Try no: ',_)
	   		print(Fore.RESET)
	   		try:
	   			user = getpass.getuser()
	   			print(Fore.YELLOW+'Hello •_•',user,'\n')
	   			usr=input("USER-NAME: ")
	   			passwd=getpass.getpass(prompt="PASSWORD: ")
	   			host ='localhost'
	   			co(host,usr,passwd)
	   			cursor=con.cursor()
	   			print(Fore.CYAN+'\nCONNECTION SUCCESSFUL\n')
	   			d=input('Enter Your Database Name: ')
	   			try:
	   				co(host,usr,passwd,d)
	   				print(f'{d} database is exsit')
	   			except:
	   					print('Database is not exsit')
	   					db=input('Do you want to create(y/n): ')
	   					if db == 'y':
	   						cursor.execute(f'create database {d}')
	   						print(f'{d} database created')
	   					else:
	   						ex()

	   			os.system('clear')
	   			break
	   		
	   		except Exception as e:
	   			print(Fore.RESET,Fore.RED)
	   			os.system('clear')
	   			print('ERORR: '+str(e))
	   			print('\nCONNECTION UNSUCCESSFUL')
	   			if _ == 3:
	   				ex()
	   			print('Try Again\n')
	   			

	co(host,usr,passwd,d)
	cursor=con.cursor()
	cursor.execute(f"create table if not exists {time}(Roll_No int primary key,Name varchar(200),Attendence varchar(200) not null default 'Absent')")
	for i in col[1:]:
		nam=c[int(i)]
		cursor.execute(f"insert into {time}(Roll_No,Name) values('{i}','{nam}')")
		con.commit()
os.system('clear')
def Help():
	print(Fore.GREEN + "=> For exit type 'exit()'\n=> To see all list of student type 'show()' ")
	print("=> To clear type 'clear()' ")
	print("=> For help type 'help()' ")
	if Q==2:
		print("=> To see attendence type 'table()' for mysql only")
Help()
print('\n************************************************************')
while True:
	
	print('\n  Present',count,'   Absent',len(col)-1-count,'     Total=',len(col)-1)
	
	print(' ')
	
	a=input(enter).title()
	print(' ')
	if l.count(a)==0:
		try:
			if a == 'Clear()':
				os.system('clear')
			elif a== 'Help()':
				info()
				eval(a)
			elif a == 'Table()':
				cursor.execute(f'select * from {time}')
				print_t(cursor)
			elif a == 'Show()':
				print(tab)
			elif a=='Exit()':
				break
			elif c[1:].count(a)>0 or col[1:].count(int(a))>0 :
				l.append(a)
				n=dict(zip(col,c))
				try:
					a=int(a)
					print(' ',n[a],'is mark present')
					count=count+1
				except:
					a=str(a)
					print(' ',a,'is mark present')
					count=count+1
				
				if Q==1:
					a=int(a)
					try:
							excel_sheet.cell(row=a+1, column=m_col).value ='Present'
							excel_sheet.cell(row=a+1, column=m_col).fill=PatternFill('solid',fgColor='00FF00')
							excel_file.save(file)
					except:
						s.write(a,x-1,'present')
						wb.save(file)
				if Q==2:
							a=int(a)
							print('mysql......')
							cursor.execute(f"update {time} set Attendence = 'Present' where Roll_No = {a}")
							con.commit()
												
			else:
				print('\nNot found')
		except:
			print('\nError found ')
	else:
		print('\nalready marked present')
print('Thank you have a nice day')
